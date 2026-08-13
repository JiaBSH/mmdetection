#!/usr/bin/env python3
"""Generate a formatted xlsx from the semantic + instance metrics summary."""

import json
import os
import sys

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


def main():
    metrics_dir = sys.argv[1] if len(sys.argv) > 1 else (
        "/data/home/scvi576/run/JiaBSH/mmdetection_para/"
        "work_dirs/run_syn_rotation/semantic_metrics"
    )
    xlsx_path = os.path.join(metrics_dir, "summary.xlsx")

    # Collect per-model results
    rows = []
    for fname in sorted(os.listdir(metrics_dir)):
        if not fname.endswith("_semseg.json"):
            continue
        with open(os.path.join(metrics_dir, fname)) as f:
            data = json.load(f)

        name = data["model_name"]
        o = data["overall"]
        inst = data.get("instance") or {}

        bbox = inst.get("coco/bbox_mAP")
        segm = inst.get("coco/segm_mAP")

        rows.append({
            "Model": name,
            "bbox_mAP": bbox,
            "segm_mAP": segm,
            "Precision": o["precision"],
            "Recall": o["recall"],
            "IoU": o["iou"],
            "F1": o["f1"],
            "TP": o["tp"],
            "FP": o["fp"],
            "FN": o["fn"],
            "TN": o["tn"],
        })

    # Sort by semantic IoU descending
    rows.sort(key=lambda r: r["IoU"], reverse=True)

    # ---- build workbook ----------------------------------------------------
    wb = openpyxl.Workbook()

    # -- Sheet 1: summary ----------------------------------------------------
    ws = wb.active
    ws.title = "Summary"

    # Styles
    header_font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_align = Alignment(horizontal="center", vertical="center")
    num_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    top1_fill = PatternFill("solid", fgColor="C6EFCE")  # green
    top2_fill = PatternFill("solid", fgColor="FFEB9C")  # yellow
    top3_fill = PatternFill("solid", fgColor="FFC7CE")  # red (low)

    # Title row
    ws.merge_cells("A1:K1")
    title_cell = ws["A1"]
    title_cell.value = "Instance Segmentation + Semantic Segmentation Metrics (run_syn_rotation)"
    title_cell.font = Font(name="Arial", bold=True, size=14)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # Sub-header
    ws.merge_cells("A2:K2")
    ws["A2"].value = (
        "Semantic metrics derived by merging instance masks (union by class). "
        "Single-class (畴区) binary segmentation on 20 test images."
    )
    ws["A2"].font = Font(name="Arial", italic=True, size=9, color="666666")
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 18

    # Headers (row 3)
    headers = [
        ("A", "Model", 50),
        ("B", "bbox_mAP", 12),
        ("C", "segm_mAP", 12),
        ("D", "Semantic\nPrecision", 14),
        ("E", "Semantic\nRecall", 14),
        ("F", "Semantic\nIoU", 14),
        ("G", "Semantic\nF1", 14),
        ("H", "TP", 12),
        ("I", "FP", 12),
        ("J", "FN", 12),
        ("K", "TN", 12),
    ]
    for col_letter, title, width in headers:
        cell = ws[f"{col_letter}3"]
        cell.value = title
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[col_letter].width = width
    ws.row_dimensions[3].height = 32

    # Data rows (starting row 4)
    for i, row in enumerate(rows):
        r = i + 4
        vals = [
            row["Model"],
            row["bbox_mAP"],
            row["segm_mAP"],
            row["Precision"],
            row["Recall"],
            row["IoU"],
            row["F1"],
            row["TP"],
            row["FP"],
            row["FN"],
            row["TN"],
        ]
        for j, val in enumerate(vals):
            cell = ws.cell(row=r, column=j + 1)
            if j == 0:
                cell.value = val
                cell.alignment = Alignment(horizontal="left", vertical="center")
            elif isinstance(val, float):
                cell.value = round(val, 4)
                cell.number_format = "0.0000"
                cell.alignment = num_align
            else:
                cell.value = val if val is not None else "N/A"
                cell.alignment = cell_align
            cell.border = thin_border

        # Highlight top 3 / bottom 3 by IoU
        if i == 0:
            fill = top1_fill
        elif i == 1:
            fill = top2_fill
        elif i == len(rows) - 1 and row["IoU"] < 0.6:
            fill = top3_fill
        else:
            fill = None

        if fill:
            for j in range(11):
                ws.cell(row=r, column=j + 1).fill = fill

    # Freeze panes
    ws.freeze_panes = "A4"

    # Auto-filter
    ws.auto_filter.ref = f"A3:K{3 + len(rows)}"

    # ---- Sheet 2: per-image detail (top model) -----------------------------
    ws2 = wb.create_sheet("Per-Image (mask2former)")

    top_model = rows[0]["Model"]
    detail_path = os.path.join(metrics_dir, f"{top_model}_semseg.json")
    if os.path.exists(detail_path):
        with open(detail_path) as f:
            detail = json.load(f)

        ws2.merge_cells("A1:G1")
        ws2["A1"].value = f"Per-Image Semantic Metrics — {top_model}"
        ws2["A1"].font = Font(name="Arial", bold=True, size=13)
        ws2.row_dimensions[1].height = 24

        detail_headers = ["Image", "Precision", "Recall", "IoU", "F1", "TP", "FP", "FN", "TN"]
        for j, h in enumerate(detail_headers):
            cell = ws2.cell(row=2, column=j + 1)
            cell.value = h
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            cell.border = thin_border
        ws2.column_dimensions["A"].width = 24
        for col in "BCDEFGHI":
            ws2.column_dimensions[col].width = 13

        for i, (img, m) in enumerate(sorted(detail["per_image"].items())):
            r = i + 3
            vals = [img, m["precision"], m["recall"], m["iou"], m["f1"],
                    m["tp"], m["fp"], m["fn"], m["tn"]]
            for j, val in enumerate(vals):
                cell = ws2.cell(row=r, column=j + 1)
                cell.value = val
                cell.alignment = cell_align if j == 0 else num_align
                cell.border = thin_border
                if j >= 1 and isinstance(val, float):
                    cell.number_format = "0.0000"

        ws2.freeze_panes = "A3"

    # ---- Sheet 3: all per-model detail -------------------------------------
    ws3 = wb.create_sheet("All Models Detail")
    ws3.merge_cells("A1:G1")
    ws3["A1"].value = "Per-Model Semantic + Instance Metrics"
    ws3["A1"].font = Font(name="Arial", bold=True, size=13)

    for j, (col_l, h, w) in enumerate(headers):
        cell = ws3.cell(row=2, column=j + 1)
        cell.value = h.replace("\n", " ")
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws3.column_dimensions[col_l].width = w
    ws3.row_dimensions[2].height = 24

    for i, row in enumerate(rows):
        r = i + 3
        vals = [
            row["Model"], row["bbox_mAP"], row["segm_mAP"],
            row["Precision"], row["Recall"], row["IoU"], row["F1"],
            row["TP"], row["FP"], row["FN"], row["TN"],
        ]
        for j, val in enumerate(vals):
            cell = ws3.cell(row=r, column=j + 1)
            if j == 0:
                cell.value = val
                cell.alignment = Alignment(horizontal="left", vertical="center")
            elif isinstance(val, float):
                cell.value = round(val, 4)
                cell.number_format = "0.0000"
                cell.alignment = num_align
            else:
                cell.value = val if val is not None else "N/A"
                cell.alignment = cell_align
            cell.border = thin_border

    ws3.freeze_panes = "A3"

    # ---- save --------------------------------------------------------------
    wb.save(xlsx_path)
    print(f"Saved {xlsx_path}")
    print(f"  Sheet 1: 'Summary' — merged table sorted by semantic IoU")
    print(f"  Sheet 2: 'Per-Image (mask2former)' — per-image breakdown for best model")
    print(f"  Sheet 3: 'All Models Detail' — full raw data")


if __name__ == "__main__":
    main()
